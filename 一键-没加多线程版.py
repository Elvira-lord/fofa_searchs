import base64
import os.path
import random
import re
import sys
from re import search
import requests
from Crypto.SelfTest.Cipher.test_CFB import file_name
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


#设置颜色
#------------------------------------------
try:
    if sys.platform=="win32":
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    green="\033[92m"     # 绿色
    red="\033[91m"       #红色
    blue="\033[94m"      #蓝色
    yellow="\033[93m"    #黄色lld
    reset = "\033[0m"     # 重置颜色
except:
    #颜色设置失败就没有颜色
    green=""
    red=""
    reset=""
    blue=""
    yellow=""
#-------------------------------------------



#读取fofa_key，读取文件配置
#-------------------------------------------
with open('config.ini','r',encoding='utf-8') as f:
    content=f.readlines()
    if len(content[4])>15:
        fofa_key=(content[4].strip())[10:-1]
    else:
        print(f'{red}请配置fofa_key!!!{reset}')
    fofa_url=content[5].strip()[10:-1]
    fofa_url=fofa_url+'/api/v1/search/all'
    fofa_email=content[6].strip()[12:-1]

    #需要额外保存的状态码
    code = content[12].strip()
    save_code = [int(i) for i in code[1:-1].strip().split(",")]
    # print(save_code)   #[200, 302, 404, 500]

    #每次查询的size
    size = int(content[14].strip().split('=')[1])
    # print(f'{size}')    #5
#-------------------------------------------



# 爬取头信息(默认)
#========================================
header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0'
}

#随机UA的列表
with open('ua.txt', 'r', encoding='utf-8') as f:
    ua_list = [line.strip() for line in f.readlines() if line.strip()]


#查询fofa余额
#=============================
def fofa_Balance():
    #接口地址
    url=content[5].strip()[10:-1]+'/api/v1/info/my'
    # print(url)
    params={
        'email': fofa_email,
        'key': fofa_key
    }
    response=requests.get(url,params=params,headers=header)
    print(f'{red}今日剩余点：{green}{response.json()['remain_free_point']}{reset}')
    return response.json()['remain_free_point']


#fofa搜索
#--------------------------------------------
def fofa_search(search='domain="baidu.com"',size=100):
    balance_count = fofa_Balance()   #今日余额

    #search 搜索内容
    #size  返回结果大小

    #base64编码查询语句
    search_base64=base64.b64encode(search.encode()).decode()

    #构造请求
    params={
        'email':fofa_email,
        'key':fofa_key,
        'qbase64':search_base64,
        'size':size
    }
    url=fofa_url+'?email='+fofa_email+'&key='+fofa_key+'&qbase64='+search_base64+'&size='+str(size)
    # print(url)

    response=requests.get(fofa_url,params=params,headers=header)
    if response.status_code==200:
        result = response.json()
        if result["error"]:
            print(f'{blue}今日余额点数:{green}{balance_count}')
            return {}
        return response.json()
    else:
        print(f"{red}fofa搜索请求失败!!!{reset}")
        return {}



#创建表存储结果
#==========================================
def create_excel(excel_path,save_code):
    if os.path.exists(excel_path):
        wb=load_workbook(excel_path)
        ws=wb.active
        if ws['A1'].value!='链接url':
            ws['A1'] = '链接url'
            ws['B1'] = 'ip'
            ws['C1'] = 'port'
            ws['D1'] = '状态'
            ws['E1'] = '标题'
            ws['F1'] = '端口类型'
            ws['G1'] = '服务'
            wb.save(excel_path)
            print('表头创建成功！！')
    else:
        wb=Workbook()  #创建一个excel表对象
        ws=wb.active   #获取当前活动的工作表
        #设置表头
        ws['A1']='链接url'
        ws['B1']='ip'
        ws['C1']='port'
        ws['D1'] = '状态'
        ws['E1'] = '标题'
        ws['F1'] = '端口类型'
        ws['G1'] = '服务'
        wb.save(excel_path)
        # print(f'文件{excel_path}已创建！！！')
    #接下来创建save_code中的表头======================================
    for code in save_code:
        if str(code) not in wb.sheetnames:
            ws_new_sheet=wb.create_sheet(str(code))
            ws_new_sheet['A1'] = '链接url'
            ws_new_sheet['B1'] = 'ip'
            ws_new_sheet['C1'] = 'port'
            ws_new_sheet['D1'] = '状态'
            ws_new_sheet['E1'] = '标题'
            ws_new_sheet['F1'] = '端口类型'
            ws_new_sheet['G1'] = '服务'
    wb.save(excel_path)




#==============================================
#常见端口和映射
non_web_ports = [
    3306, 1433, 1521, 5432, 27017, 6379,  # 数据库
    22, 23, 3389, 5900, 5901,  # 远程管理
    21, 20, 445, 139,  # 文件传输
    25, 110, 143, 465, 993, 995,  # 邮件服务
    53, 67, 68,  # DNS/DHCP
    873, 11211, 2181, 9092, 5672, 15672  # 其他服务
]
# 端口到服务的映射
port_service_map = {
    # ===== 数据库 =====
    3306: "MySQL",1433: "MSSQL",1521: "Oracle",5432: "PostgreSQL",27017: "MongoDB",6379: "Redis",
    9200: "Elasticsearch",9300: "Elasticsearch-Cluster",
    # ===== 远程管理 =====
    22: "SSH",23: "Telnet",3389: "RDP",5900: "VNC",5901: "VNC-1",
    # ===== 文件传输 =====
    21: "FTP",20: "FTP-Data",445: "SMB",139: "NetBIOS",
    # ===== 邮件服务 =====
    25: "SMTP",110: "POP3",143: "IMAP",465: "SMTPS",993: "IMAPS",995: "POP3S",
    # ===== DNS/DHCP =====
    53: "DNS",67: "DHCP-Server",68: "DHCP-Client",
    # ===== 其他服务 =====
    873: "rsync",11211: "Memcached",2181: "Zookeeper",9092: "Kafka",5672: "RabbitMQ",
    15672: "RabbitMQ-Management",
}


daiLiChi=[]
#代理池
#如果前面 没有写协议的话会自动添加 http:// ，所以代理池中最好都是加上协议的
with open('代理池.txt','r',encoding='utf-8') as f:
    for i in f.readlines():
        proxy=i.strip()
        if proxy and not proxy.startswith(('http://','https://','socks5://')):
            proxy='http://'+proxy
        daiLiChi.append(proxy.strip())
# print(daiLiChi)    #这个是代理池

#存储
#将存储的列表进行存储
def save_result(list,last_row,excel_path):
    '''
    #接下来 将得到的数据写入excel
    #看情况吧，多半要写一个 def 用来搞这个
    url   链接
    ip
    port   端口
    statusCode  状态码
    title  标题
    ip_type  默认web    -----这个要改
    server   服务
    ['http://m.yuntu.baidu.com', '182.61.200.83', '80', 200, ['页面不存在_百度搜索'], 'web', 'Apache']
    '''
    print('#',list)
    wb=load_workbook(excel_path)
    ws=wb.active
    while ws[f'A{last_row}'].value is not None:
        last_row+=1
    ws[f'A{last_row}']=list[0]  #url
    ws[f'B{last_row}']=list[1]  #ip
    ws[f'C{last_row}']=list[2]  #端口
    ws[f'D{last_row}']=list[3]  #状态码
    if len(list[4])==0:
        ws[f'E{last_row}']='!!!未找到标题'
    else:
        ws[f'E{last_row}']=list[4][0]  #标题
    ws[f'F{last_row}']=list[5]  #端口类型
    ws[f'G{last_row}']=list[6]  #服务
    wb.save(excel_path)

#=========
#这里写存储其他status_code的存储
def save_other_status_code(list,status_code,save_code_last_row,excel_path):
    wb=load_workbook(excel_path)
    ws=wb[str(status_code)]
    while ws[f'A{save_code_last_row[status_code]}'].value is not None:
        save_code_last_row[status_code]+=1
    ws[f'A{save_code_last_row[status_code]}'] = list[0]  # url
    ws[f'B{save_code_last_row[status_code]}'] = list[1]  # ip
    ws[f'C{save_code_last_row[status_code]}'] = list[2]  # 端口
    ws[f'D{save_code_last_row[status_code]}'] = list[3]  # 状态码
    if len(list[4]) == 0:
        ws[f'E{save_code_last_row[status_code]}'] = '!!!未找到标题'
    else:
        ws[f'E{save_code_last_row[status_code]}'] = list[4][0]  # 标题
    ws[f'F{save_code_last_row[status_code]}'] = list[5]  # 端口类型
    ws[f'G{save_code_last_row[status_code]}'] = list[6]  # 服务
    wb.save(excel_path)



#探活
#=======================================================================
def request_url(ips,ip_type,code_200,last_row,save_code_last_row,save_code,excel_path):
    #调用前面的随机ua
    ua=random.choice(ua_list)
    header_ua={
        'User-Agent':ua
    }

    if ips[0].startswith('http'):
        # ===========================================================================
        # 这里加一个代理池，如果获取失败，就使用代理池，还失败就没招了
        try:
            print(f'{green}[*]{ips[0]}{reset}',end='   ')
            zhilian=False  #先直连，如果直连失败就代理，代理失败就没招
            title=[]
            statusCode=0
            server='未知'
            port = ips[2]  # 端口
            ip = ips[1]  # ip
            url = ips[0]  # 链接

            try:
                response = requests.get(ips[0], timeout=5, headers=header_ua)
                response.encoding = 'utf-8'
                statusCode=response.status_code     #状态码
                title=re.findall('<title>(.*?)</title>', response.text)    #标题
                if len(title)==0:
                    title.append('！！！无标题')
                server=response.headers.get('Server','未知')    #服务
                zhilian=True
                print(f'{green}[直连]成功{reset}')
            except Exception as e:
                print(f'{yellow}切换代理{reset}',end='   ')
            if not zhilian and len(daiLiChi):
                for i,proxy in enumerate(daiLiChi,1):
                    try:
                        proxys={
                            'http':proxy,
                            'https':proxy
                        }
                        print(f'{yellow}尝试代理{red}[{i}/{len(daiLiChi)}]{reset}',end='  ')
                        response = requests.get(ips[0], timeout=5, headers=header_ua,proxies=proxys)
                        response.encoding = 'utf-8'
                        statusCode=response.status_code
                        title=re.findall('<title>(.*?)</title>', response.text)
                        if len(title) == 0:
                            title.append('！！！无标题')
                        server=response.headers.get('Server','未知')
                        print(f'{green}{ips[0]}  [代理]成功{reset}')
                    except Exception as e:
                        print()
                        print(f'{red}----------------------  {proxy}{reset}  -->{red}代理连接失败{reset}')
        except Exception as e:
            print(f'{red}[&]{ips[0]} 连接失败{reset}')
    else:
        if int(ips[2]) in non_web_ports:
            ip_type=port_service_map[int(ips[2])]
        else:
            ip_type='未知'

    '''
    #接下来 将得到的数据写入excel
    url   链接
    ip
    port   端口
    statusCode  状态码
    title  标题
    ip_type  默认web    -----这个要改
    server   服务         
    '''

    result=[ips[0],ips[1],ips[2],statusCode,title,ip_type,server]
    save_result(result,last_row[0],excel_path)
    last_row[0] += 1  #每次加1
    if statusCode==200:
        code_200[0]+=1
    if statusCode in save_code:
        save_other_status_code(result,statusCode,save_code_last_row,excel_path)
        save_code_last_row[statusCode]+=1


def ping(save_code,results_list,excel_path):
    code_200 = [0]  # 记录200的数量

    last_row = [2]  # 主表的最后一行
    save_code_last_row={}
    for i in save_code:   #其他code的最后一行
        save_code_last_row[i]=2
    #print(save_code_last_row)  #测试  {200: 2, 302: 2, 404: 2, 500: 2}

    for ips in results_list :
        ip_type = 'web'
        if not ips[0].startswith('http') and int(ips[2]) not in non_web_ports:
            url = 'http://' + ips[0]
            ips[0] = url
        if ips[0].startswith('http'):
            pass
        request_url(ips,ip_type,code_200,last_row,save_code_last_row,save_code,excel_path)
    print(f'{green}[$$$]status.code=200的数量为：{red}{code_200[0]}{reset}')
    return code_200[0]


def get_file_name(search_name,is_add_200count):
    with open('config.ini', 'r', encoding='utf-8') as f:
        content = f.readlines()
        #保存的文件名
        if content[9].strip()[17:]=='True':    #原来这里是True
            # 如果为True说明要保存在一个文件以内
            file_name = re.search(r'[\'"](.*?)[\'"]', content[10].strip()).group(1)  # 提取捕获组为文件名
            excel_path = f'result/{file_name}'
        else:   #当为False的时候，就将搜索中第一个的内容用为文件名
            # 这个就用第一个引号内包裹的内容为文件名吧，如果文件名存在就在后面加上序号，然后最后再给文件名最后加上文件中200的数量
            file_name=re.search(r"['\"](.*?)[\"]",search_name).group(1)
            excel_path=f'result/{file_name}.xlsx'
            is_add_200count[0]=1
            if os.path.exists(excel_path):
                i=1
                while True:
                    excel_path=f'result/{file_name}_{i}.xlsx'
                    if not os.path.exists(excel_path):
                        break
                    i+=1
    return excel_path

def 批量搜索():
    with open('批量搜索.txt', 'r', encoding='utf-8') as f:
        search_list = [line.strip() for line in f.readlines() if line.strip()]
        # print(search_list)  #['domain="baidu.com"', 'doamin="360"']
        return search_list

#主程序,方便多次调用
def main(search_name):
    #如果是分别保存的模式就在名字后面加上code=200的数量
    is_add_200count=[0]
    excel_path=get_file_name(search_name,is_add_200count)
    print(is_add_200count)  #
    # print(fofa_key)
    # print(fofa_url)
    # print(fofa_email)
    # 搜索后的是 json 格式
    # print(fofa_search())
    # ========================================
    search_result = fofa_search(search_name, size)  # 搜索结果
    # print('========================',search_result)
    if len(search_result) > 0:
        print(f'{red}正在查询语句:{green}{search_result['query']}{reset}')  # 本次查询语句
        # print(f'{red}本次查询需要积分:{green}{search_result['required_fpoints']}{reset}')  #本次查询需要积分
        print(f'{red}本次消耗积分:{green}{search_result['consumed_fpoint']}{reset}')  # 本次消耗积分
    else:
        print(f"{blue}今日余额不足！！！")

    create_excel(excel_path,save_code)  # 执行一下

    # 存储结果
    # ==============================================
    results_list = search_result['results']
    # print(results_list)

    code_200=ping(save_code,results_list,excel_path)
    # print(code_200)

    #给这些文件名后面加上code=200的数量
    if is_add_200count[0]==1:
        f_name=re.search(r"[/](.*)",excel_path).group(1)[:-5]
        #print(f_name)  #result
        rename_excel_path=f'result/{f_name}_({code_200}).xlsx'
        if os.path.exists(rename_excel_path):
            i=1
            while True:
                rename_excel_path=f'result/{f_name}_{i}_({code_200}).xlsx'
                if not os.path.exists(rename_excel_path):
                    break
                i+=1
        os.rename(excel_path,rename_excel_path)


search_list=批量搜索()
if len(search_list)==0:
    while True:
        print(f"{yellow}~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~{reset}")
        search_name=input(f"{red}输入搜索内容(fofa):{reset}")
        if search_name=='':
            print(f"{red}退出")
            break
        try:
            main(search_name)
        except:
            print(f"{red}{search_name}搜索失败！！！")
else:
    print(f"{blue}批量搜索开始......^_^{reset}")
    for search_name in search_list:
        try:
            main(search_name)
        except:
            print(f"{red}{search_name}搜索失败！！！")
    print(f"{blue}批量搜索结束......^_^")

